import argparse
import math
import sys
import time
from dataclasses import dataclass, field
import requests
from tabulate import tabulate
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from github import Github
from dotenv import load_dotenv
load_dotenv()

anchor_brands = [
    "walmart", "walmart supercenter", "walmart neighborhood market", "target", "costco", "sams club", "sam's club",
    "nugget markets", "safeway", "kroger", "lucky", "sprouts farmers market", "sprouts",
    "raley's", "raleys", "bel-air", "bel air", "belair", "bristol farms", "bristol", "gelsons", "gelson's",
    "erewhon", "erewhon market", "trader joes", "trader joe's", "vons", "vons supermarket", "vons market", "albertsons",
    "albertson's", "albertson", "ralphs", "ralphs supermarket", "ralphs market", "smart & final", "smart and final", 
    "smart & final extra", "smart and final extra", "publix", "whole foods", "whole foods market", "wholefoods", 
    "wholefoods market", "foodmaxx", "99 ranch market", "walgreens", "cvs", "big lots", "planet fitness", "gold's gym", "crunch fitness", "lifetime fitness", "california family fitness", "24 hour fitness",

    "home depot", "the home depot", "lowe's", "lowes", "best buy", "macy's", "macys", "nordstrom", "nordstrom rack", "bloomingdale's", 
    "bloomingdales", "dicks sporting goods", "dick's sporting", "dick's sporting goods", "dollar tree", "save mart", "grocery outlet",
    "winco", "bevmo", "bevmo!", "harbor freight", "homegoods", "home goods", "ross", "burlington", "tj maxx", "william-sonoma",
    "petco", "h&m", "marshalls", "kohl's","ace hardware", "office depot", "big 5", "rite aid", "jcpenney", "nike", "adidas"
]
subbrand_noise = [
    "gasoline", "gas", "gas station", "fuel", "fuel station", "pharmacy", "garden center", "garden centre", "car wash", "tire center",
    "tire centre", "auto center", "auto centre", "optical", "portrait studio", "photo center", "deli", "bakery", "jewelry", "money center",
    "vision center", "furniture", "furniture gallery", "express", "convenience store",
]

plaza_radius_mi = 0.18
min_anchors_per_plaza = 1
search_radius_km = 5
min_other_tenants = 1

overpass_mirrors = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.private.coffee/api/interpreter",
    "https://overpass.openstreetmap.ru/api/interpreter",
]
headers = {"User-Agent": "retailfinder/1.0"}


@dataclass
class Store:
    name:str
    lat:float
    lng: float
    brand: str = ""
    address: str= ""
    city: str = ""
    is_anchor_store: bool = False

@dataclass
class Plaza:
    anchors: list=field(default_factory=list)
    tenants: list=field(default_factory=list)
    mall_name: str = ""
    mall_address: str = ""
    county: str=""

    @property
    def all_stores(self):
        return self.anchors + self.tenants
    
    @property
    def center(self):
        pts = self.all_stores
        lats = [s.lat for s in pts]
        lngs = [s.lng for s in pts]
        return sum(lats) / len(lats), sum(lngs) / len(lngs)
    
    @property
    def label(self):
        return self.mall_name if self.mall_name else "Unnamed Retail Center"

    @property
    def display_address(self):
        if self.mall_address:
            return self.mall_address
        for s in self.all_stores:
            if is_valid_address(s.address):
                return s.address
        return "-"
    @property
    def display_city(self):
        for s in self.all_stores:
            if s.city:
                return s.city
        addr = self.display_address
        if addr and addr != "-":
            return addr.strip().split()[-1]
        return "-"
    
    @property
    def anchor_names(self):
        return ", ".join(sorted(set(s.name for s in self.anchors)))
    
    @property
    def tenant_names(self):
        return ", ".join(sorted(set(s.name for s in self.tenants)))


def geocode_city(city:str) -> tuple[float,float, str]:
    parts = [p.strip() for p in city.split(",")]
    city_name = parts[0]
    state = parts[1].strip() if len(parts) >1 else ""

    query = f"""
[out:json][timeout:30];
(
    node["place"~"city|town"]["name"~"^{city_name}$",i];
);
out 5;
""".strip()
    
    elements = run_overpass(query)

    if not elements:
        raise ValueError(
            f"City not found: '{city}'. \n"
            "Use 'City, State' format such as 'Roseville, CA'."
        )
    chosen = elements[0]
    if state:
        for el in elements:
            tags = el.get("tags",{})
            st = tags.get("addr:state", "") or tags.get("is_in:state_code","")
            if state.upper() in st.upper():
                chosen = el
                break
    tags = chosen.get("tags",{})
    name = tags.get("name",city_name)
    state_tag = tags.get("addr:state") or tags.get("is_in:state_code") or state
    display = f"{name}, {state_tag}" if state_tag else name
    return chosen["lat"], chosen["lon"], display

def build_store_query(lat:float,lng:float,radius_km:float) -> str:
    radius_m = int(radius_km * 1000)
    return f"""
[out:json][timeout:40];
(
    node["shop"](around:{radius_m},{lat},{lng});
    node["brand"](around:{radius_m},{lat},{lng});
    way["shop"](around:{radius_m},{lat},{lng});
    way["brand"](around:{radius_m},{lat},{lng});
);
out center tags;
""".strip()

def build_mall_query(lat:float,lng:float,radius_km:float) -> str:
    radius_m = int(radius_km * 1000)
    return f"""
[out:json][timeout:30];
(
    node["shop"="mall"](around:{radius_m},{lat},{lng});
    way["shop"="mall"](around:{radius_m},{lat},{lng});
    relation["shop"="mall"](around:{radius_m},{lat},{lng});
    way["landuse"="retail"]["name"](around:{radius_m},{lat},{lng});
);
out center tags;
""".strip()

def build_county_query(lat:float, lng: float) -> str:
    return f"""
[out:json][timeout:15];
is_in({lat}, {lng});
out tags;
""".strip()

def run_overpass(query: str, retries: int = 2) -> list:
    last_error = None
    for mirror in overpass_mirrors:
        for attempt in range(retries):
            try:
                resp = requests.post(mirror, data={"data": query}, headers=headers, timeout=90)
                if resp.status_code == 400:
                    print(f"\n  [Overpass 400] Query:\n{query}\n  Response: {resp.text[:300]}")
                    resp.raise_for_status()
                if resp.status_code in (502, 503, 504):
                    wait = 3 * (attempt + 1)
                    print(f"  [warn] {mirror} returned {resp.status_code}, waiting {wait}s...")
                    time.sleep(wait)
                    last_error = RuntimeError(f"HTTP {resp.status_code} from {mirror}")
                    continue
                resp.raise_for_status()
                return resp.json().get("elements", [])
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                wait = 3 * (attempt + 1)
                print(f"  [warn] {mirror} timed out (attempt {attempt+1}/{retries}), waiting {wait}s...")
                last_error = e
                time.sleep(wait)
                continue
            except requests.exceptions.HTTPError as e:
                raise RuntimeError(f"Overpass query failed: {e}") from e
        print(f"  [warn] {mirror} exhausted, moving to next mirror...")
    raise RuntimeError(f"All Overpass mirrors failed. Last error: {last_error}")

def is_anchor(name:str) -> bool:
    if not name:
        return False
    n = name.lower().strip()
    for brand in anchor_brands:
        pattern = r'\b' + re.escape(brand) + r'\b'
        if re.search(pattern, n):
            return True
    return False

def is_subbrand(name:str) -> bool:
    if not name:
        return False
    n = name.lower().strip()
    matches_anchor = any(re.search(r'\b' + re.escape(b) + r'\b', n) for b in anchor_brands)
    matches_noise  = any(re.search(r'\b' + re.escape(noise) + r'\b', n) for noise in subbrand_noise)
    return matches_anchor and matches_noise

street_types = {
    "street", "st", "dr", "drive", "ln", "lane", "blvd", "boulevard", "way","rd","road", "ave", "avenue", "ct", "court", "pl", "place","cir", "circle",
    "hwy", "highway", "pkwy", "parkway", "trail", "trl", "loop", "run", "expy", "expressway"
}
def is_valid_address(address:str) -> bool:
    if not address:
        return False
    tokens = address.strip().split()
    if len(tokens) < 4:
        return False
    last = tokens[-1].lower().rstrip(".")
    if last in street_types:
        return False
    if not tokens[0].isdigit():
        return False
    return True

def extract_stores(elements:list) -> list[Store]:
    stores = []
    seen = set()

    for el in elements:
        tags = el.get("tags",{})
        name = tags.get("name") or tags.get("brand") or ""
        if not name:
            continue
        if is_subbrand(name):
            continue


        lat = el.get("lat") or (el.get("center") or {}).get("lat")
        lng = el.get("lon") or (el.get("center") or {}).get("lon")
        if lat is None or lng is None:
            continue

        key = (round(lat,4), round(lng,4))
        if key in seen:
            continue
        seen.add(key)

        addr_city = tags.get("addr:city","")
        address_parts = [
            tags.get("addr:housenumber",""),
            tags.get("addr:street",""),
            addr_city,
        ]
        address = " ".join(p for p in address_parts if p).strip()

        stores.append(Store(
            name=name,
            lat=lat,
            lng=lng,
            brand=tags.get("brand",""),
            address=address,
            city = addr_city,
            is_anchor_store= is_anchor(name),
        ))

    return stores

def haversine_m(lat1,lng1,lat2,lng2) -> float:
    r = 6_371_000
    dlat = math.radians(lat2-lat1)
    dlng = math.radians(lng2-lng1)
    a = (math.sin(dlat/2) ** 2
         +math.cos(math.radians(lat1))
         *math.cos(math.radians(lat2))
         *math.sin(dlng/2) ** 2
         )
    return r*2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

def build_plazas (stores: list[Store], radius_miles: float, min_other_tenants: int) -> list[Plaza]:
    radius_m = radius_miles * 1609.34
    anchors = [s for s in stores if s.is_anchor_store]

    used = set()
    anchor_groups: list[list[Store]] = []
    for i, a in enumerate(anchors):
        if i in used:
            continue
        group = [a]
        used.add(i)
        for j, b in enumerate(anchors):
            if j in used:
                continue
            if all(haversine_m(m.lat, m.lng, b.lat, b.lng) <= radius_m for m in group):
                group.append(b)
                used.add(j)
        anchor_groups.append(group)

    plazas = []
    for plaza_anchors in anchor_groups:
        clat = sum(a.lat for a in plaza_anchors) / len(plaza_anchors)
        clng = sum(a.lng for a in plaza_anchors) / len(plaza_anchors)

        anchor_coords = {(round(a.lat, 4), round(a.lng, 4)) for a in plaza_anchors}

        tenants = []
        for s in stores:
            coord = (round(s.lat, 4), round(s.lng, 4))
            if coord in anchor_coords:
                continue
            d = haversine_m(clat, clng, s.lat, s.lng)
            if d <= radius_m:
                tenants.append(s)

        seen_coords = set()
        unique_tenants = []
        for t in tenants:
            coord = (round(t.lat, 4), round(t.lng, 4))
            if coord not in seen_coords:
                seen_coords.add(coord)
                unique_tenants.append(t)

        non_anchor_tenants = [t for t in unique_tenants if not t.is_anchor_store]
        extra_anchor_tenants = [t for t in unique_tenants if t.is_anchor_store]
        plaza_anchors = plaza_anchors + extra_anchor_tenants

        if len(non_anchor_tenants) >= min_other_tenants:
            plazas.append(Plaza(anchors=plaza_anchors, tenants=non_anchor_tenants))

    return plazas

def deduplicate_plaza_stores(plazas: list[Plaza]) -> list[Plaza]:
    def dedupe_list(items:list[Store]) -> list[Store]:
        seen_brands: dict[str,Store] = {}
        for store in items:
            key = store.name.lower().strip()
            if key not in seen_brands:
                seen_brands[key] = store
            else:
                if is_valid_address(store.address) and not is_valid_address(seen_brands[key].address):
                    seen_brands[key] = store
        return list(seen_brands.values())
    for plaza in plazas:
        plaza.anchors = dedupe_list(plaza.anchors)
        plaza.tenants = dedupe_list(plaza.tenants)

    plazas_sorted = sorted(plazas, key = lambda p: len(p.all_stores), reverse = True)

    global_seen: set[tuple] = set()
    final_plazas = []
    for plaza in plazas_sorted:
        new_anchors, new_tenants =[], []
        for a in plaza.anchors:
            coord = (round(a.lat,4), round(a.lng,4))
            if coord not in global_seen:
                global_seen.add(coord)
                new_anchors.append(a)
        for t in plaza.tenants:
            coord = (round(t.lat,4), round(t.lng,4))
            if coord not in global_seen:
                global_seen.add(coord)
                new_tenants.append(t)
        plaza.anchors = new_anchors
        plaza.tenants = new_tenants

        if len(plaza.anchors) >= 1 and len(plaza.tenants) >= min_other_tenants:
            final_plazas.append(plaza)
    return final_plazas

def merge_same_name_plazas(plazas: list[Plaza]) -> list[Plaza]:
    named: dict[str,Plaza] = {}
    unnamed: list[Plaza] = []

    for plaza in plazas:
        if not plaza.mall_name:
            unnamed.append(plaza)
        else:
            key = plaza.mall_name.strip().lower()
            if key not in named:
                named[key] = plaza
            else:
                existing = named[key]
                existing.anchors.extend(plaza.anchors)
                existing.tenants.extend(plaza.tenants)
                if not is_valid_address(existing.mall_address) and is_valid_address(plaza.mall_address):
                    existing.mall_address = plaza.mall_address

    merged = list(named.values()) + unnamed

    def dedupe_list(items: list[Store]) -> list[Store]:
        seen_brands: dict[str, Store] = {}
        for store in items:
            key = store.name.lower().strip()
            if key not in seen_brands:
                seen_brands[key] = store
            else:
                if is_valid_address(store.address) and not is_valid_address(seen_brands[key].address):
                    seen_brands[key] = store
        return list(seen_brands.values())
    
    for plaza in merged:
        plaza.anchors = dedupe_list(plaza.anchors)
        plaza.tenants = dedupe_list(plaza.tenants)
    
    return sorted(
        [p for p in merged if len(p.anchors) >= 1 and len(p.tenants) >= min_other_tenants],
        key = lambda p: len(p.all_stores),
        reverse = True,
    )

def attach_mall_names(plazas: list[Plaza], mall_elements:list)-> None:
    malls = []
    for el in mall_elements:
        tags = el.get("tags", {})
        name = tags.get("name")
        if not name:
            continue
        lat = el.get("lat") or (el.get("center") or {}).get("lat")
        lng = el.get("lon") or (el.get("center") or {}).get("lon")
        if not lat or not lng:
            continue
        address_parts = [
            tags.get("addr:housenumber", ""),
            tags.get("addr:street", ""),
            tags.get("addr:city"),
        ]
        address = " ".join(p for p in address_parts if p).strip()
        malls.append((name, lat, lng, address))


    for plaza in plazas:
        clat,clng = plaza.center
        best_name,best_addr,best_dist = None, "", float("inf")
        for name, mlat, mlng,address in malls:
            d = haversine_m(clat,clng,mlat,mlng)
            if d < best_dist and d <= 600:
                best_dist = d
                best_name = name
                best_addr = address
        plaza.mall_name = best_name or ""
        plaza.mall_address = best_addr or ""


def lookup_county(lat:float,lng:float) -> str:
    try: 
        elements = run_overpass(build_county_query(lat,lng))
        for el in elements:
            tags = el.get("tags", {})
            if (tags.get("admin_level") == "6" and tags.get("boundary") == "administrative"):
                name = tags.get("name", "")
                return name  
    except Exception:
        pass
    return "-"
def attach_counties(plazas:list[Plaza]) -> None:
    cache: dict[str, str] = {}
    for i, plaza in enumerate(plazas):
        clat,clng = plaza.center
        cache_key = f"{round(clat,2)}, {round(clng,2)}"
        if cache_key not in cache:
            print(f"  [county] looking up cluster {i + 1}/{len(plazas)}...", end="\r")
            cache[cache_key] = lookup_county(clat,clng)
            time.sleep(0.5)
        plaza.county = cache[cache_key]
    print(" " * 50, end="\r")




def js_escape(s:str) -> str:
    return (s.replace("\\", "\\\\")
             .replace("'","\\'")
             .replace("\n"," ")
             .replace("\r",""))
def generate_map(plazas: list[Plaza], city_display: str, radius_miles: float,all_stores: list[Store], output_path = None) -> None:
    if output_path is None:
        slug = city_display.lower().replace(", ", "-").replace(" ", "-")
        output_path = f"{slug}.html"
    if not plazas:
        return

    # Center the map on the average of all cluster centers
    all_lats = [p.center[0] for p in plazas]
    all_lngs = [p.center[1] for p in plazas]
    map_lat = sum(all_lats) / len(all_lats)
    map_lng = sum(all_lngs) / len(all_lngs)
    base_radius_m = radius_miles * 1609.34

    coord_to_color = {}
    for p in plazas:
        color = "#e74c3c" if p.mall_name else "#3498db"
        for s in p.all_stores:
            coord_to_color[(round(s.lat,4), round(s.lng,4))] = (color,p.label)
    store_js = []
    for s in all_stores:
        sname = js_escape(s.name)
        saddr = js_escape(s.address if s.address else "-")
        key = (round(s.lat,4), round(s.lng,4))

        if key in coord_to_color:
            color, plaza_label = coord_to_color[key]
            if s.is_anchor_store:
                dot_color = "#f39c12"
            else:
                dot_color = color
        else:
            color = "#27ae60"
            plaza_label = "Standalone / Not in Plaza"
            dot_color = "#27ae60"


        plaza_label = js_escape(plaza_label)
        store_js.append(
            f"  addStore({s.lat}, {s.lng}, '{sname}', '{saddr}', '{dot_color}', '{plaza_label}', {str(s.is_anchor_store).lower()});"
        )
    stores_code = "\n".join(store_js)

    plaza_js = []
    for p in plazas:
        clat, clng = p.center
        max_dist = max(
            (haversine_m(clat,clng,s.lat,s.lng) for s in p.all_stores),
            default = base_radius_m,
        )
        plaza_radius_m = max(max_dist*1.15,base_radius_m)
        anchors_list = js_escape(p.anchor_names)
        tenants_list = js_escape(p.tenant_names)
        label       = js_escape(p.label)
        address     = js_escape(p.display_address)
        county      = js_escape(p.county)
        city        = js_escape(p.display_city)
        color = "#e74c3c" if p.mall_name else "#3498db"

        plaza_js.append(
            f"  addPlaza({clat}, {clng}, {plaza_radius_m:.1f}, '{label}', "
            f"'{address}', '{county}', '{city}', {len(p.anchors)}, '{anchors_list}', "
            f"{len(p.tenants)}, '{tenants_list}', '{color}');"
        )


    plazas_code = "\n".join(plaza_js)

    html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset='utf-8' />
  <title>Retail Centers – {city_display}</title>
  <meta name='viewport' content='width=device-width, initial-scale=1.0'>
  <link rel='stylesheet' href='https://unpkg.com/leaflet@1.9.4/dist/leaflet.css' />
  <script src='https://unpkg.com/leaflet@1.9.4/dist/leaflet.js'></script>
  <style>
    body {{ margin: 0; font-family: sans-serif; }}
    #map {{ height: 100vh; width: 100%; }}
    .legend {{
      background: white; padding: 10px 14px; border-radius: 6px;
      box-shadow: 0 1px 5px rgba(0,0,0,0.3); line-height: 1.8;
      font-size: 13px;
    }}
    .legend-dot {{ display: inline-block; width: 12px; height: 12px;
                   border-radius: 50%; margin-right: 6px; }}
  </style>
</head>
<body>
<div id='map'></div>
<script>
  var map = L.map('map').setView([{map_lat}, {map_lng}], 12);

  L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png', {{
    attribution: '© <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a>',
    maxZoom: 19
  }}).addTo(map);

  function addPlaza(lat, lng, radius_m, name, address, county, city, numAnchors, anchors, numTenants, tenants, color) {{
    L.circle([lat, lng], {{
      radius: radius_m,
      color: color,
      fillColor: color,
      fillOpacity: 0.08,
      weight: 2
    }}).addTo(map);

    var popup = '<b>' + name + '</b><br>'
      + (address !== '-' ? address + '<br>' : '')
      + city + (county !== '-' ? ', ' + county + ' County' : '') + '<br>'
      + '<br><b>' + numAnchors + ' anchor(s):</b><br>'
      + anchors.split(', ').join('<br>')
      + '<br><br><b>' + numTenants + ' other tenant(s):</b><br>'
      + tenants.split(', ').join('<br>');

    L.circleMarker([lat, lng], {{
      radius: 9,
      color: '#fff',
      fillColor: '#8e44ad',
      fillOpacity: 1,
      weight: 2
    }}).bindPopup(popup).bindTooltip(name, {{permanent: false, direction: 'top'}}).addTo(map);
  }}

  function addStore(lat, lng, name, address, color, plaza, isAnchor) {{
    var size = isAnchor ? 11 : 9;
    var icon = L.divIcon({{
      className: '',
      html: '<div style="position:absolute;width:' + size + 'px;height:' + size + 'px;border-radius:50%;background:' + color + ';border:2px solid white;box-shadow:0 1px 3px rgba(0,0,0,0.4);top:50%;left:50%;transform:translate(-50%,-50%)' + (isAnchor ? ';outline:2px solid rgba(0,0,0,0.25)' : '') + '"></div>',
      iconSize: [size, size],
      iconAnchor: [size/2, size/2]
    }});
    L.marker([lat, lng], {{icon: icon}})
      .bindPopup('<b>' + name + '</b>' + (isAnchor ? '⚓' : '')+ '<br>' 
        + '<i>' + plaza + '</i><br>'
        + address)
      .bindTooltip(name, {{direction: 'top'}})
      .addTo(map);
  }}

  {plazas_code}
  {stores_code}

  var legend = L.control({{position: 'bottomright'}});
  legend.onAdd = function() {{
    var div = L.DomUtil.create('div', 'legend');
    div.innerHTML = '<b>Retail Centers – {city_display}</b><br>'
      + '<span class="legend-dot" style="background:#e74c3c"></span> Named plaza<br>'
      + '<span class="legend-dot" style="background:#3498db"></span> Unnamed plaza<br>'
      + '<span class="legend-dot" style="background:#8e44ad"></span> Plaza center point<br>'
      + '<span class="legend-dot" style="background:#f39c12"></span> Anchor store (in plaza)<br>'
      + '<span class="legend-dot" style="background:#27ae60"></span> Standalone store<br>'
      + 'Circle = plaza footprint &nbsp; ⚓ = anchor';
    return div;
  }};
  legend.addTo(map);
</script>
</body>
</html>"""

    with open(output_path, "w") as f:
        f.write(html)
    print(f"\n  Map saved → {output_path}  (open in any browser)")
    return output_path

def print_results(plazas: list[Plaza], city_display: str, args) -> None:
    print(f"\n{'='*60}")
    print(f"  Major retail centers near {city_display}")
    print(f"  Criteria: {args.min_anchors}+ anchors, {args.min_tenants} + other tenant(s) within {args.radius} mile radius")
    print(f"{'='*60}\n")

    if not plazas:
        print("  No qualifying retail clusters found.")
        print("  Try lowering --min-anchors or increasing --radius.\n")
        return

    # Sort by county then city alphabetically
    sorted_plazas = sorted(
        plazas,
        key=lambda p: (
            p.county.lower() if p.county and p.county != "-" else "zzz",
            p.display_city.lower() if p.display_city and p.display_city != "-" else "zzz",
        )
    )

    print(f"  Found {len(sorted_plazas)} major retail center(s):\n")

    table_rows = [
        [
            p.label,
            args.city.split(",")[1].strip() if "," in args.city else "-",
            p.county,
            p.display_city,
            p.display_address,
            len(p.anchors),
            p.anchor_names,
            len(p.tenants),
            p.tenant_names
        ]
        for p in sorted_plazas
    ]

    print(tabulate(
        table_rows,
        headers=["Plaza / Property Name", "State", "County", "City", "Address", "# Anchors", "Anchor Names", "# Tenants", "Other Tenants"],
        tablefmt="simple",
    ))
    print()
 

def export_to_excel(plazas, city_display, args, map_url, output_path = os.path.expanduser("~/Desktop/text.xlsx")):
    if os.path.exists(output_path):
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
    else: 
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    tab_name = city_display.replace(",","").strip()[:31]

    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(title=tab_name)


    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "test"
    title_cell.font = Font(name = "Arial", bold = True, size = 16, color = "FFFFFF")
    title_cell.fill = PatternFill("solid", start_color = "2C3E50")
    title_cell.alignment = Alignment(horizontal = "center", vertical = "center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:I2")
    map_cell = ws["A2"]
    
    map_cell.hyperlink = map_url
    map_cell.value = "View Interactive Map (Live)" if map_url.startswith("http") else "View Interactive Map (Local)"

    map_cell.font = Font(name="Arial", bold = True, size = 11, color = "FFFFFF",underline="single")
    map_cell.fill = PatternFill("solid", start_color="2980B9")
    map_cell.alignment = Alignment(horizontal = "center", vertical = "center")
    ws.row_dimensions[2].height = 24

    headers = ["Plaza / Property Name", "State", "County", "City", "Address", "# Anchors", "Anchor Names", "# Tenants", "Other Tenants"]
    header_fill = PatternFill("solid", start_color = "34495E")
    header_font = Font(name = "Arial", bold = True, size = 11, color = "FFFFFF")
    thin = Side(style = "thin", color = "CCCCCC")
    border = Border(left=thin, right=thin, top=thin,bottom = thin)

    for col, header in enumerate(headers,1):
        cell = ws.cell(row=3, column = col, value = header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border=border
    ws.row_dimensions[3].height=20

    sorted_plazas = sorted(
        plazas,
        key=lambda p: (
            p.county.lower() if p.county and p.county != "-" else "zzz",
            p.display_city.lower() if p.display_city and p.display_city != "-" else "zzz",
        )
    )
    state = args.city.split(",")[1].strip() if "," in args.city else "-"
    row_fills = [
        PatternFill("solid", start_color = "FDFEFE"),
        PatternFill("solid", start_color = "EBF5FB"),
    ]

    for i, plaza in enumerate(sorted_plazas):
        row = i+4
        fill = row_fills[i % 2]
        data = [
            plaza.label,
            state,
            plaza.county,
            plaza.display_city,
            plaza.display_address,
            len(plaza.anchors),
            plaza.anchor_names,
            len(plaza.tenants),
            plaza.tenant_names,
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column = col, value = value)
            cell.font = Font(name="Arial", size = 10)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical = "center", wrap_text =(col in (7,9)))

            if col == 5 and value and value != "-":
                maps_url = f"https://www.google.com/maps/search/?api=1&query={requests.utils.quote(value)}"
                cell.value = value
                cell.hyperlink = maps_url
                cell.font = Font(name="Arial", size=10, color = "0563C1", underline = "single")
        ws.row_dimensions[row].height = 18

    for col, width in enumerate([35,8,18,18,38,10,40,10,55], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.freeze_panes = "A4"
    wb.save(output_path)
    print(f"\n Excel Saved -> {output_path}")


def upload_map_to_github(html_path:str) -> str:
    token = os.getenv("GITHUB_TOKEN")
    username = os.getenv("GITHUB_USERNAME")
    repo_name = os.getenv("GITHUB_REPO", "retail-maps")

    if not token or not username:
        print("  [warn] github_token or github_username not set in .env")
        return os.path.abspath(html_path)
    
    try:
        import github as gh
        g = gh.Github(auth=gh.Auth.Token(token))
        user = g.get_user()

        try:
            repo = user.get_repo(repo_name)
            print(f"  [github] Found repo: {repo_name}")
        except Exception:
            print(f"  [github] Creating repo: {repo_name}")
            repo = user.create_repo(
                repo_name,
                description = "Retail center maps",
                auto_init = True,
            )
            time.sleep(2)
        filename = os.path.basename(html_path)

        with open(html_path, "r", encoding="utf-8") as f:
            content = f.read()
        
        try:
            existing = repo.get_contents(filename)
            repo.update_file(
                filename,
                f"Update {filename}",
                content,
                existing.sha
            )
            print(f"  [github] Updated {filename}")
        except Exception:
            repo.create_file(
                filename,
                f"Add {filename}",
                content,
            )
            print(f"  [github] Created {filename}")

        try:
            repo.enable_pages(source={"branch": "main", "path": "/"})
            print(f"  [github] GitHub Pages enabled")
        except Exception:
            pass 

        url = f"https://{username}.github.io/{repo_name}/{filename}"
        print(f"  [github] Map live at: {url}")
        return url

    except Exception as e:
        print(f"  [warn] GitHub upload failed: {e}")
        print(f"  [warn] Falling back to local path.")
        return os.path.abspath(html_path)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
 
def main():
    parser = argparse.ArgumentParser(
        description="Find major retail centers in a city using OpenStreetMap."
    )
    parser.add_argument("city", help='City to search, e.g. "Roseville, CA"')
    parser.add_argument(
        "--radius", type=float, default=plaza_radius_mi,
        help=f"Cluster radius in miles (default: {plaza_radius_mi})"
    )
    parser.add_argument(
        "--min-anchors", type=int, default=min_anchors_per_plaza,
        help=f"Minimum anchor stores per cluster (default: {min_anchors_per_plaza})"
    )
    parser.add_argument(
        "--min-tenants", type=int, default=min_other_tenants,
        help=f"Minimum other (non-anchor) tenants required to confirm a real plaza (default: {min_other_tenants})"
    )
    parser.add_argument(
        "--search-km", type=float, default=search_radius_km,
        help=f"Search radius from city center in km (default: {search_radius_km})"
    )
    args = parser.parse_args()
 
    print(f"\n  Searching for major retail centers in: {args.city}")
 
    print("  [1/5] Geocoding city...")
    lat, lng, display = geocode_city(args.city)
    print(f"        → {display} ({lat:.4f}, {lng:.4f})")
    time.sleep(1)
 
    print("  [2/5] Querying OpenStreetMap for stores (may take 10–20s)...")
    store_elements = run_overpass(build_store_query(lat, lng, args.search_km))
    print(f"        → {len(store_elements)} raw elements returned")
 
    print("  [3/5] Querying for mall/retail area names...")
    try:
        mall_elements = run_overpass(build_mall_query(lat, lng, args.search_km))
        print(f"        → {len(mall_elements)} mall/retail areas found")
    except RuntimeError as e:
        print(f"  [warn] Mall name lookup failed ({e}). Centers will show as 'Unnamed'.")
        mall_elements = []
 
    stores = extract_stores(store_elements)
    n_anchors = sum(1 for s in stores if s.is_anchor_store)
    print(f"\n  Total shops identified: {len(stores)}  ({n_anchors} are anchors)")
    if n_anchors == 0:
        print("  No anchor stores found. OSM data may be sparse for this area.")
        sys.exit(0)
    if not stores:
        print("  No anchor stores found. OSM data may be sparse for this area.")
        sys.exit(0)
 
    print("  [4/5] Building plazas around each anchor and gathering tenants...")
    plazas = build_plazas(stores, args.radius, args.min_tenants)
    attach_mall_names(plazas, mall_elements)
    plazas = deduplicate_plaza_stores(plazas)
    plazas = merge_same_name_plazas(plazas)
 
    print(f"  [5/5] Looking up counties for {len(plazas)} plazas...")
    attach_counties(plazas)
 
    print_results(plazas, display, args)
    map_path = generate_map(plazas, display, args.radius, stores)
    print("\n  Uploading to GitHub...")
    map_url = upload_map_to_github(map_path)
    export_to_excel(plazas, display, args, map_url)
 
 
if __name__ == "__main__":
    main()